from selenium import webdriver
from webdriver_manager.chrome import ChromeDriverManager
import openpyxl
def get_chromedriver_path():
    driver_path = ChromeDriverManager().install()
    return  driver_path

def readExlData(path,sheetName):
    try:
        wb = openpyxl.load_workbook(path)
        sheet = wb.get_sheet_by_name(sheetName)
        return sheet, wb
    except Exception as e:
        return e

def getExlRowCount(sheet):
    try:
        # sheet = readExlData(path,sheetName)
        return (sheet.max_row)
    except Exception as e:
        return e
def getExlColCount(sheet):
    try:
        # sheet = readExlData(path,sheetName)
        return (sheet.max_column)
    except Exception as e:
        return e
def getColNames(sheet):
    try:
        col_name = [i.value for i in sheet[1]]
        return col_name
    except Exception as e:
        return e

def getRowVal(sheet,rowNo):
    # To Get Specific Row Values
    try:
        # sheet = readExlData(path,sheetName)
        columnNames = getColNames(sheet)
        col = getExlColCount(sheet)
        data = {}
        for i in range(1,col+1):
            rowVal = sheet.cell(row=rowNo,column=i).value
            print(rowVal)
            data.update({columnNames[i-1]:rowVal})
        return data
    except Exception as e:
        return e

def updateExlData(sheet,workbook,path,data,testCase,colNo=5):
    # This is currently set to update order_number column with colNo = 5
    try:
        rowCount = getExlRowCount(sheet)
        for i in range(2,rowCount+1):
            if sheet.cell(row=i,column=1).value == testCase:
                sheet.cell(row=i,column=colNo).value = data
                workbook.save(path)
                return True
        # sheet.cell(row=rowNo, column=colNo).value = data
        # workbook.save(path)
        # workbook.close()
    except Exception as e:
        return e

# fileName = 'F:\\Projects\\rf_DataDriven\\testdata\\TestData.xlsx'
# sheetname = 'loginTest'
# sheet, wb = readExlData(fileName,sheetname)
# print(sheet)
# print(wb)
# # print(getRowVal(fileName,sheetname,2))
# # print(getColNames(sheet))
# # print(next(getRowVal(fileName,sheetname,1)))
# b = updateExlData(sheet,wb,fileName,2,'FAILED')
# print(b)